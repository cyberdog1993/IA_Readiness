from __future__ import annotations

import csv
import datetime as dt
import io
import subprocess
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


BASE_DIR = Path("/Users/cyberdog/Documents/IA_Readiness")
BACKUP_DIR = BASE_DIR / "backups"
EXPORT_DIR = BASE_DIR / "exports"
OUTPUT_FILE = EXPORT_DIR / "IA-readiness_respuestas_actuales_2026-07-17.xlsx"


def latest_backup() -> Path | None:
    backups = sorted(BACKUP_DIR.glob("consultores_it_*.sql"), key=lambda item: item.stat().st_mtime)
    return backups[-1] if backups else None


def fetch_leads_csv() -> str:
    command = [
        "ssh",
        "deploy@192.168.93.27",
        "sudo -n docker exec consultores-it-postgres psql -U consultores -d consultores_it -Atc \"COPY (SELECT * FROM leads ORDER BY created_at DESC) TO STDOUT WITH CSV HEADER\"",
    ]
    return subprocess.check_output(command, text=True)


def parse_value(column: str, raw: str):
    if raw == "":
        return None

    integer_fields = {
        "repetitive_process_count",
        "manual_hours_weekly",
        "process_documentation_level",
        "digital_system_usage",
        "excel_dependency",
        "system_integration_level",
        "manual_report_generation",
        "key_person_dependency",
        "automation_interest",
        "maturity_score",
    }
    boolean_fields = {"has_kpis", "privacy_consent"}
    datetime_fields = {"consulting_requested_at", "created_at", "updated_at", "deleted_at"}

    if column in integer_fields and raw.isdigit():
        return int(raw)
    if column in boolean_fields:
        return raw.lower() in {"t", "true", "1", "yes"}
    if column in datetime_fields:
        try:
            return dt.datetime.fromisoformat(raw)
        except ValueError:
            return raw
    return raw


def style_header(row_cells):
    header_fill = PatternFill("solid", fgColor="0F172A")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in row_cells:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")


def autosize_columns(ws, max_width: int = 45):
    for column_cells in ws.columns:
        values = [cell.value for cell in column_cells if cell.value is not None]
        if not values:
            continue
        width = min(max(len(str(value)) for value in values) + 2, max_width)
        ws.column_dimensions[get_column_letter(column_cells[0].column)].width = max(12, width)


def main() -> Path:
    EXPORT_DIR.mkdir(parents=True, exist_ok=True)

    raw_csv = fetch_leads_csv()
    rows = list(csv.DictReader(io.StringIO(raw_csv)))
    backup = latest_backup()

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Respuestas"

    headers = list(rows[0].keys()) if rows else []
    if headers:
        sheet.append(headers)
        style_header(sheet[1])

    for row in rows:
        sheet.append([parse_value(column, row.get(column, "")) for column in headers])

    sheet.freeze_panes = "A2"
    if headers:
        sheet.auto_filter.ref = sheet.dimensions
    autosize_columns(sheet)

    summary = workbook.create_sheet("Resumen")
    summary.append(["Métrica", "Valor"])
    style_header(summary[1])

    scores = [int(row["maturity_score"]) for row in rows if row.get("maturity_score", "").isdigit()]
    summary_rows = [
        ("Total de respuestas", len(rows)),
        ("Promedio de madurez", round(sum(scores) / len(scores), 1) if scores else None),
        ("Máxima madurez", max(scores) if scores else None),
        ("Mínima madurez", min(scores) if scores else None),
        ("Respaldo previo", backup.name if backup else "N/A"),
        ("Generado en", dt.datetime.now()),
    ]
    for item in summary_rows:
        summary.append(list(item))

    summary.freeze_panes = "A2"
    summary.column_dimensions["A"].width = 24
    summary.column_dimensions["B"].width = 42

    rollback = workbook.create_sheet("Rollback")
    rollback.append(["Archivo", "Ruta"])
    style_header(rollback[1])
    rollback.append(["Respaldo SQL", str(backup) if backup else "N/A"])
    rollback.append(["Uso", "Restaurar antes de cambios si algo falla"])
    rollback.column_dimensions["A"].width = 18
    rollback.column_dimensions["B"].width = 80

    workbook.save(OUTPUT_FILE)

    # Quick validation
    check = load_workbook(OUTPUT_FILE, read_only=True)
    assert "Respuestas" in check.sheetnames
    assert "Resumen" in check.sheetnames

    return OUTPUT_FILE


if __name__ == "__main__":
    print(main())
